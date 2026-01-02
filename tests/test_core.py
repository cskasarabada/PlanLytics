import pytest
from io import BytesIO

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None

from app.core import composer, simulate, optimize, constraints
from app.export import excel_pack, oracle_icm


def test_plan_lifecycle():
    plan = composer.suggest_plan({"sales_goal": 100000, "num_reps": 5, "commission_rate": 0.1})
    plan_id = plan["plan_id"]

    # ensure details stored
    assert composer.PLAN_STORE[plan_id]["quota_per_rep"] == 20000

    perf = {"alice": 50000}
    result = simulate.run_simulation(plan_id, perf)
    assert result["payouts"]["alice"] == 5000.0

    optimize.optimize_plan(plan_id, {"target_rate": 0.2})
    result = simulate.run_simulation(plan_id, perf)
    assert result["payouts"]["alice"] == 10000.0

    payload = oracle_icm.export_plan_to_icm(plan_id)
    assert payload["details"]["commission_rate"] == 0.2

    workbook_bytes = excel_pack.generate_workbook(plan_id)
    assert isinstance(workbook_bytes, (bytes, bytearray))
    if load_workbook:
        wb = load_workbook(BytesIO(workbook_bytes))
        assert wb.active["B1"].value == plan_id


def test_validate_plan_rejects_invalid():
    with pytest.raises(ValueError):
        constraints.validate_plan({"quota_per_rep": -1, "commission_rate": 0.1})
    with pytest.raises(ValueError):
        constraints.validate_plan({"quota_per_rep": 100, "commission_rate": 0})
