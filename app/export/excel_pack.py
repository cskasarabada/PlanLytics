"""Excel export helpers."""
from io import BytesIO
from typing import Dict

try:  # Optional dependency
    from openpyxl import Workbook
except Exception:  # pragma: no cover - fallback path
    Workbook = None  # type: ignore

from ..core import composer


def generate_workbook(plan_id: str) -> bytes:
    """Create a simple workbook for the given plan.

    If :mod:`openpyxl` is available, an actual XLSX file is produced. Otherwise
    a CSV-like representation is returned so callers still receive useful
    content.
    """
    details: Dict | None = composer.PLAN_STORE.get(plan_id)

    if Workbook is None:  # Fallback to CSV bytes
        lines = [f"Plan ID,{plan_id}"]
        if details:
            for k, v in details.items():
                lines.append(f"{k},{v}")
        return "\n".join(lines).encode()

    wb = Workbook()
    ws = wb.active
    ws.title = "Plan"
    ws["A1"] = "Plan ID"
    ws["B1"] = plan_id

    if details:
        row = 2
        for key, value in details.items():
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
            row += 1

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
