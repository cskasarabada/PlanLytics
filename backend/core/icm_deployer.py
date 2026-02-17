# core/icm_deployer.py
"""
Bridge module that invokes ICM Optimizer managers to deploy the generated
Excel workbook to Oracle Fusion ICM via REST APIs.

The manager classes are bundled under backend/icm_optimizer/ — no external
sys.path manipulation is needed.
"""
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, Optional, List

import pandas as pd

from ..icm_optimizer.utils.api_client import APIClient
from ..icm_optimizer.config.config_manager import ConfigManager
from ..icm_optimizer.core.rate_dimension import RateDimensionManager
from ..icm_optimizer.core.rate_table import RateTableManager
from ..icm_optimizer.core.expression import ExpressionManager
from ..icm_optimizer.core.perf_measure import PerformanceMeasureManager
from ..icm_optimizer.core.plan_component import PlanComponentManager
from ..icm_optimizer.core.comp_plan import CompensationPlanManager

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Detailed deployment log capture
# ---------------------------------------------------------------------------

class _DeploymentLog:
    """Captures every API request/response during deployment for audit trail."""

    def __init__(self, workbook_name: str, org_id: int, approval_tag: str = ""):
        self.workbook_name = workbook_name
        self.org_id = org_id
        self.approval_tag = approval_tag
        self.started_at = datetime.utcnow().isoformat() + "Z"
        self.current_step = ""
        self.entries: List[Dict[str, Any]] = []
        self._seq = 0

    def set_step(self, step_name: str):
        self.current_step = step_name

    def record(
        self,
        method: str,
        endpoint: str,
        url: str,
        status_code: int,
        request_payload: Any = None,
        response_body: Any = None,
    ):
        self._seq += 1
        # Truncate large response bodies (keep first 2000 chars for display)
        resp_summary = response_body
        if isinstance(response_body, dict):
            resp_summary = _truncate_response(response_body)
        elif isinstance(response_body, str) and len(response_body) > 2000:
            resp_summary = response_body[:2000] + "...(truncated)"

        self.entries.append({
            "seq": self._seq,
            "timestamp": datetime.utcnow().isoformat() + "Z",
            "step": self.current_step,
            "method": method,
            "endpoint": endpoint,
            "url": url,
            "status_code": status_code,
            "request_payload": request_payload,
            "response_body": resp_summary,
            "success": 200 <= status_code < 300,
            "action": _classify_action(method, endpoint, status_code),
        })

    def to_dict(self) -> Dict[str, Any]:
        return {
            "workbook_name": self.workbook_name,
            "org_id": self.org_id,
            "approval_tag": self.approval_tag,
            "started_at": self.started_at,
            "finished_at": datetime.utcnow().isoformat() + "Z",
            "total_requests": len(self.entries),
            "successful_requests": sum(1 for e in self.entries if e["success"]),
            "failed_requests": sum(1 for e in self.entries if not e["success"]),
            "entries": self.entries,
        }


def _truncate_response(resp: dict, max_items: int = 5) -> dict:
    """Shrink large API responses for log readability."""
    out = {}
    for k, v in resp.items():
        if k == "links":
            out[k] = f"[{len(v)} links]" if isinstance(v, list) else v
        elif k == "items" and isinstance(v, list) and len(v) > max_items:
            out[k] = v[:max_items] + [f"...({len(v) - max_items} more)"]
        elif isinstance(v, str) and len(v) > 500:
            out[k] = v[:500] + "...(truncated)"
        else:
            out[k] = v
    return out


def _classify_action(method: str, endpoint: str, status_code: int) -> str:
    """Derive a human-readable action label from method + endpoint."""
    # Extract the resource name from the endpoint
    resource = endpoint.rstrip("/").split("/")[-1].split("?")[0]
    # Clean up camelCase / snake_case
    if resource.startswith("child"):
        parts = endpoint.rstrip("/").split("/")
        resource = parts[-1] if len(parts) > 1 else resource

    if method == "GET":
        if "?q=" in endpoint:
            return f"Lookup {resource}"
        return f"Fetch {resource}"
    elif method == "POST":
        if status_code == 201:
            return f"Create {resource}"
        elif status_code == 200:
            return f"Assign {resource}"
        else:
            return f"Create {resource} (failed)"
    elif method == "PATCH":
        return f"Update {resource}"
    return f"{method} {resource}"


class _LoggingAPIClient:
    """Wraps APIClient to intercept and log every request/response."""

    def __init__(self, api_client: APIClient, deploy_log: _DeploymentLog):
        self._client = api_client
        self._log = deploy_log
        # Expose same attributes as APIClient for any direct access
        self.base_url = api_client.base_url
        self.api_path = api_client.api_path
        self.username = api_client.username
        self.password = api_client.password
        self.logger = api_client.logger

    def _build_url(self, endpoint: str) -> str:
        return self._client._build_url(endpoint)

    def get(self, endpoint: str):
        url = self._client._build_url(endpoint)
        response, status_code = self._client.get(endpoint)
        self._log.record("GET", endpoint, url, status_code,
                         request_payload=None, response_body=response)
        return response, status_code

    def post(self, endpoint: str, data: dict):
        url = self._client._build_url(endpoint)
        response, status_code = self._client.post(endpoint, data)
        self._log.record("POST", endpoint, url, status_code,
                         request_payload=data, response_body=response)
        return response, status_code

    def patch(self, endpoint: str, data: dict):
        url = self._client._build_url(endpoint)
        response, status_code = self._client.patch(endpoint, data)
        self._log.record("PATCH", endpoint, url, status_code,
                         request_payload=data, response_body=response)
        return response, status_code


# ---------------------------------------------------------------------------
# Config proxy & helpers (unchanged)
# ---------------------------------------------------------------------------

class _DirectConfigProxy:
    """Lightweight stand-in for ConfigManager when using direct credentials.

    All six ICM Optimizer manager classes call
    ``self.config_manager.get('organization')`` in __init__ to read org_id.
    This proxy satisfies that interface without needing a YAML file.
    """

    def __init__(self, org_id: int):
        self._config = {"organization": {"org_id": org_id}}

    def get(self, section: str, key: str = None, default=None):
        if key is None:
            return self._config.get(section, default)
        return self._config.get(section, {}).get(key, default)

    def get_section(self, section: str) -> Dict[str, Any]:
        return self._config.get(section, {})


def _read_org_id_from_workbook(excel_path: Path) -> int:
    """Extract OrgId from the workbook's Config sheet.

    Returns 0 if not found (caller should resolve via API).
    """
    try:
        cfg = pd.read_excel(excel_path, sheet_name="Config")
        row = cfg.loc[cfg["Key"] == "OrgId"]
        if not row.empty:
            val = int(row.iloc[0]["Value"])
            if val != 0:
                return val
    except Exception:
        pass
    return 0


def _lookup_org_id_from_api(api_client: APIClient) -> int:
    """Look up OrgId from the Oracle instance by querying existing compensation plans.

    Falls back to 0 if the lookup fails.
    """
    try:
        response, status_code = api_client.get(
            "/compensationPlans?limit=1&fields=OrgId"
        )
        if status_code == 200 and response.get("items"):
            org_id = response["items"][0].get("OrgId", 0)
            if org_id:
                logger.info("Auto-detected OrgId %d from Oracle API", org_id)
                return int(org_id)
    except Exception as e:
        logger.warning("Failed to auto-detect OrgId from API: %s", e)
    return 0


def _normalize_workbook_for_managers(excel_path: Path) -> Path:
    """Create a copy of the workbook with columns renamed for ICM Optimizer managers.

    The ICM transformer produces human-readable columns (e.g. "Expression Detail Type")
    while the ICM Optimizer managers expect camelCase (e.g. "ExpressionDetailType").
    This bridges the gap without modifying either side.

    Returns path to the normalized workbook (temp file next to original).
    """
    import shutil
    import openpyxl

    # Sheet renames (transformer sheet name -> manager expected sheet name)
    SHEET_RENAMES: Dict[str, str] = {
        "Performance Measure": "Performance Measures",
    }

    # Column renames per sheet (our column -> manager expected column)
    # Keys use the ORIGINAL sheet name (before sheet rename)
    RENAMES: Dict[str, Dict[str, str]] = {
        "Expression": {
            "Expression Detail Type": "ExpressionDetailType",
            "Basic Attributes Group": "BasicAttributesGroup",
            "Basic Attribute Name": "BasicAttributeName",
            "Measure Name": "MeasureName",
            "Measure Result Attribute": "MeasureResultAttribute",
            "Expression Operator": "ExpressionOperator",
            "Constant Value": "ConstantValue",
            "Plan Component Name": "PlanComponentName",
            "Plan Component Result Attribute": "PlanComponentResultAttribute",
        },
        "Rate Table": {
            "Rate Table Type": "RateTableType",
        },
        "Rate Table Rates": {
            "Rate Value": "RateValue",
        },
        "Performance Measure": {
            "CreditCategoryName": "Credit Category Name",
            "StartDate": "Start Date",
            "EndDate": "End Date",
        },
        "Plan Components": {
            "StartDate": "Start Date",
            "EndDate": "End Date",
        },
    }

    normalized = excel_path.parent / f"_deploy_{excel_path.name}"
    shutil.copy2(excel_path, normalized)

    wb = openpyxl.load_workbook(normalized)

    # Rename columns first (using original sheet names)
    for sheet_name, col_map in RENAMES.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=1, column=col_idx).value
            if cell_val in col_map:
                ws.cell(row=1, column=col_idx).value = col_map[cell_val]

    # Rename sheets (e.g. "Performance Measure" -> "Performance Measures")
    for old_name, new_name in SHEET_RENAMES.items():
        if old_name in wb.sheetnames:
            wb[old_name].title = new_name

    wb.save(normalized)
    wb.close()
    return normalized


# ---------------------------------------------------------------------------
# Preview & validate (unchanged)
# ---------------------------------------------------------------------------

def preview_deployment(excel_path: Path) -> Dict[str, Any]:
    """
    Read the ICM workbook and return a summary of objects that will be created.
    No API calls are made -- this is a local-only preview.
    """
    preview: Dict[str, Any] = {"success": True, "objects": {}, "total_objects": 0}

    try:
        xls = pd.ExcelFile(excel_path)
    except Exception as e:
        return {"success": False, "error": f"Cannot read workbook: {e}"}

    sheet_map = {
        "Credit Categories": "credit_categories",
        "Rate Dimension": "rate_dimensions",
        "Rate Table": "rate_tables",
        "Rate Table Rates": "rate_table_rates",
        "Expression": "expressions",
        "Performance Measure": "performance_measures",
        "Performance Goals": "performance_goals",
        "Plan Components": "plan_components",
        "Compensation Plans": "compensation_plans",
        "Calculation Settings": "calculation_settings",
        "Scorecards": "scorecards",
    }

    total = 0
    for sheet_name, key in sheet_map.items():
        if sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            rows = df.to_dict(orient="records")
            # Extract primary identifier for each row
            name_col = next(
                (c for c in df.columns if "name" in c.lower()),
                df.columns[0] if len(df.columns) > 0 else None,
            )
            names = [str(r.get(name_col, "")) for r in rows] if name_col else []
            preview["objects"][key] = {
                "count": len(rows),
                "names": names[:20],  # Cap at 20 for display
            }
            total += len(rows)

    preview["total_objects"] = total
    preview["sheets_found"] = xls.sheet_names
    return preview


def validate_credentials(
    base_url: str,
    username: str,
    password: str,
) -> Dict[str, Any]:
    """Test Oracle API credentials without deploying anything."""
    try:
        client = APIClient(base_url=base_url, username=username, password=password)
        response, status_code = client.get(
            "/incentiveCompensationPerformanceMeasures?limit=1"
        )
        if status_code == 200:
            return {"success": True, "message": "API credentials validated"}
        return {"success": False, "message": f"API returned status {status_code}"}
    except Exception as e:
        return {"success": False, "message": f"Connection failed: {e}"}


# ---------------------------------------------------------------------------
# Post-deployment configured workbook generator
# ---------------------------------------------------------------------------

def generate_configured_workbook(
    original_excel: Path, deploy_log: _DeploymentLog
) -> Path:
    """Create a post-deployment workbook enriched with Oracle object IDs.

    Reads the original pre-deployment workbook, extracts Oracle-assigned IDs
    from the deploy log entries, and writes a new workbook with an additional
    "Oracle ID" column per sheet plus a "Deployment Summary" sheet.

    Returns the path to the configured workbook.
    """
    import openpyxl
    from copy import copy

    # Extract Oracle IDs from deploy log by parsing POST/201 responses
    oracle_ids: Dict[str, Dict[str, Dict[str, Any]]] = {
        # step -> { object_name -> { field: value, ... } }
    }

    for entry in deploy_log.entries:
        if entry.get("method") != "POST" or entry.get("status_code") != 201:
            continue
        resp = entry.get("response_body") or {}
        step = entry.get("step", "")
        name = resp.get("Name", "")
        if not name:
            continue

        ids: Dict[str, Any] = {}
        # Collect all Oracle ID fields from the response
        for key, val in resp.items():
            if key.endswith("Id") and isinstance(val, (int, float)) and val:
                ids[key] = int(val)
        if ids:
            oracle_ids.setdefault(step, {})[name] = ids

    # Also extract IDs from GET lookups (for objects that already existed)
    for entry in deploy_log.entries:
        if entry.get("method") != "GET" or "?q=" not in entry.get("endpoint", ""):
            continue
        resp = entry.get("response_body") or {}
        items = resp.get("items", [])
        if not items or not isinstance(items, list):
            continue
        step = entry.get("step", "")
        first = items[0]
        name = first.get("Name", "")
        if not name:
            continue
        # Don't overwrite if we already have a POST result
        if step in oracle_ids and name in oracle_ids[step]:
            continue

        ids = {}
        for key, val in first.items():
            if key.endswith("Id") and isinstance(val, (int, float)) and val:
                ids[key] = int(val)
        if ids:
            oracle_ids.setdefault(step, {})[name] = ids

    # Map deploy log step names to workbook sheet names
    step_to_sheet = {
        "Rate Dimensions": ("Rate Dimension", "Rate Dimension Name", "RateDimensionId"),
        "Rate Tables": ("Rate Table", "Rate Table Name", "RateTableId"),
        "Expressions": ("Expression", "Expression Name", "ExpressionId"),
        "Performance Measures": ("Performance Measure", "Name", "PerformanceMeasureId"),
        "Plan Components": ("Plan Components", "Plan Component Name", "PlanComponentId"),
        "Compensation Plans": ("Compensation Plans", "Name", "CompensationPlanId"),
    }

    # Copy original workbook and enrich with Oracle IDs
    configured_path = original_excel.parent / original_excel.name.replace(
        "icm_", "icm_configured_"
    )
    import shutil
    shutil.copy2(original_excel, configured_path)

    wb = openpyxl.load_workbook(configured_path)

    for step_name, (sheet_name, name_col, id_field) in step_to_sheet.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        step_ids = oracle_ids.get(step_name, {})
        if not step_ids:
            continue

        # Find the name column index
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        name_col_idx = None
        for idx, h in enumerate(headers, 1):
            if h == name_col:
                name_col_idx = idx
                break
        if name_col_idx is None:
            continue

        # Add Oracle ID column header
        id_col_idx = ws.max_column + 1
        ws.cell(row=1, column=id_col_idx, value="Oracle ID")
        # Style the header
        header_cell = ws.cell(row=1, column=id_col_idx)
        # Copy style from the first header cell if available
        src_cell = ws.cell(row=1, column=1)
        if src_cell.font:
            header_cell.font = copy(src_cell.font)
        if src_cell.fill:
            header_cell.fill = copy(src_cell.fill)
        if src_cell.alignment:
            header_cell.alignment = copy(src_cell.alignment)

        # Add a second column for Status
        status_col_idx = id_col_idx + 1
        ws.cell(row=1, column=status_col_idx, value="Deploy Status")
        status_header = ws.cell(row=1, column=status_col_idx)
        if src_cell.font:
            status_header.font = copy(src_cell.font)
        if src_cell.fill:
            status_header.fill = copy(src_cell.fill)

        # Fill in Oracle IDs for each row
        for row_idx in range(2, ws.max_row + 1):
            obj_name = ws.cell(row=row_idx, column=name_col_idx).value
            if obj_name and str(obj_name).strip() in step_ids:
                ids = step_ids[str(obj_name).strip()]
                oracle_id = ids.get(id_field, "")
                if oracle_id:
                    id_cell = ws.cell(row=row_idx, column=id_col_idx, value=oracle_id)
                    id_cell.number_format = '0'  # Prevent scientific notation
                    ws.cell(row=row_idx, column=status_col_idx, value="✓ Deployed")
                else:
                    ws.cell(row=row_idx, column=status_col_idx, value="✓ Found")
            else:
                ws.cell(row=row_idx, column=status_col_idx, value="— Not deployed")

    # Also enrich Credit Categories sheet
    if "Credit Categories" in wb.sheetnames:
        ws = wb["Credit Categories"]
        cc_ids = oracle_ids.get("Performance Measures", {})
        # Credit category IDs come from PM step entries
        cc_name_ids = {}
        for entry in deploy_log.entries:
            resp = entry.get("response_body") or {}
            if "CreditCategoryId" in resp and "Name" in resp:
                cc_name_ids[resp["Name"]] = resp["CreditCategoryId"]

        if cc_name_ids:
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            name_col_idx = None
            for idx, h in enumerate(headers, 1):
                if h in ("Name", "Credit Category Name"):
                    name_col_idx = idx
                    break
            if name_col_idx:
                id_col_idx = ws.max_column + 1
                ws.cell(row=1, column=id_col_idx, value="Oracle CreditCategoryId")
                for row_idx in range(2, ws.max_row + 1):
                    cc_name = ws.cell(row=row_idx, column=name_col_idx).value
                    if cc_name and str(cc_name).strip() in cc_name_ids:
                        cell = ws.cell(row=row_idx, column=id_col_idx,
                                       value=cc_name_ids[str(cc_name).strip()])
                        cell.number_format = '0'

    # Add Deployment Summary sheet
    if "Deployment Summary" in wb.sheetnames:
        del wb["Deployment Summary"]
    ws_summary = wb.create_sheet("Deployment Summary", 0)

    # Header styling
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="left", vertical="center")

    # Summary header
    summary_headers = ["Field", "Value"]
    for col, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    log_dict = deploy_log.to_dict()
    summary_data = [
        ("Workbook", log_dict.get("workbook_name", "")),
        ("OrgId", log_dict.get("org_id", "")),
        ("Approval", log_dict.get("approval_tag", "")),
        ("Deployment Started", log_dict.get("started_at", "")),
        ("Deployment Finished", log_dict.get("finished_at", "")),
        ("Total API Requests", log_dict.get("total_requests", 0)),
        ("Successful Requests", log_dict.get("successful_requests", 0)),
        ("Failed Requests", log_dict.get("failed_requests", 0)),
        ("", ""),  # Blank row
        ("— Object IDs Created —", ""),
    ]

    # Add all created object IDs
    for step_name, objects in oracle_ids.items():
        for obj_name, ids in objects.items():
            primary_id_field = step_to_sheet.get(step_name, (None, None, None))[2]
            primary_id = ids.get(primary_id_field, next(iter(ids.values()), ""))
            summary_data.append((f"{step_name}: {obj_name}", primary_id))

    for row_idx, (field, value) in enumerate(summary_data, 2):
        ws_summary.cell(row=row_idx, column=1, value=field)
        val_cell = ws_summary.cell(row=row_idx, column=2, value=value)
        if isinstance(value, (int, float)) and value > 100000:
            val_cell.number_format = '0'

    # Auto-size columns
    ws_summary.column_dimensions["A"].width = 40
    ws_summary.column_dimensions["B"].width = 30

    wb.save(configured_path)
    wb.close()

    logger.info("Post-deployment configured workbook: %s", configured_path)
    return configured_path


# ---------------------------------------------------------------------------
# Main deployment function
# ---------------------------------------------------------------------------

def deploy_to_oracle_icm(
    excel_path: Path,
    config_path: Optional[Path] = None,
    dry_run: bool = False,
    *,
    base_url: str = "",
    username: str = "",
    password: str = "",
    org_id: int = 0,
    approval_tag: str = "",
) -> Dict[str, Any]:
    """
    Deploy ICM configuration from Excel workbook to Oracle ICM via ICM Optimizer.

    Credentials can come from either config_path (YAML) or direct parameters.

    Args:
        excel_path: Path to the generated ICM workbook
        config_path: Path to ICM Optimizer config.yaml (optional if direct creds)
        dry_run: If True, return preview without deploying
        base_url: Oracle Fusion base URL (direct credential mode)
        username: API username (direct credential mode)
        password: API password (direct credential mode)
        org_id: Explicit OrgId (0 = auto-detect from workbook/API)
        approval_tag: Approval reference (e.g. analysis_id + approval date)

    Returns:
        Dict with deployment results per step + detailed_log
    """
    results: Dict[str, Any] = {"steps": [], "success": True}

    # Dry run: preview what would be created
    if dry_run:
        preview = preview_deployment(excel_path)
        results["mode"] = "dry_run"
        results["message"] = "Preview only -- no objects created in Oracle ICM"
        results["preview"] = preview
        logger.info("Dry run: returning deployment preview")
        return results

    # Initialize API client from config file or direct credentials
    try:
        if base_url and username and password:
            # Direct credential mode -- no config file needed
            raw_api_client = APIClient(
                base_url=base_url,
                username=username,
                password=password,
            )
            # OrgId resolution order: explicit param -> workbook -> API auto-detect
            if org_id == 0:
                org_id = _read_org_id_from_workbook(excel_path)
            if org_id == 0:
                org_id = _lookup_org_id_from_api(raw_api_client)
            if org_id == 0:
                results["success"] = False
                results["error"] = (
                    "Could not determine OrgId. Set it in the workbook Config "
                    "sheet or ensure the Oracle instance has at least one "
                    "compensation plan."
                )
                return results
            logger.info("Using OrgId: %d", org_id)
            config_manager = _DirectConfigProxy(org_id)
        elif config_path:
            config_manager = ConfigManager(str(config_path))
            api_config = config_manager.get_section("api")
            raw_api_client = APIClient(
                base_url=api_config["base_url"],
                username=api_config["username"],
                password=api_config["password"],
            )
        else:
            results["success"] = False
            results["error"] = "Provide either config_path or direct credentials"
            return results
    except Exception as e:
        results["success"] = False
        results["error"] = f"Failed to initialize API client: {e}"
        logger.error(results["error"])
        return results

    # Create the deployment log and logging API client wrapper
    deploy_log = _DeploymentLog(
        workbook_name=excel_path.name,
        org_id=org_id,
        approval_tag=approval_tag,
    )
    api_client = _LoggingAPIClient(raw_api_client, deploy_log)

    # Normalize workbook columns for ICM Optimizer managers
    normalized_path = _normalize_workbook_for_managers(excel_path)
    excel_str = str(normalized_path)
    log_file = "objects_created.log"

    # Build manager kwargs
    def _mgr_kwargs():
        kw = {"log_file": log_file, "excel_path": excel_str}
        return kw

    # Create shared PerformanceMeasureManager for PlanComponentManager dependency
    pm_manager = PerformanceMeasureManager(
        api_client, config_manager, **_mgr_kwargs()
    )

    # Shared ExpressionManager for expression creation + validation
    expr_manager = ExpressionManager(
        api_client, config_manager, **_mgr_kwargs()
    )

    def _validate_and_fix_expressions() -> bool:
        """Post-creation step: check all expressions, fix INVALID ones.

        Oracle expressions start with Status=INVALID.  They auto-transition
        to VALID when ExpressionDetails are correctly set.  If any are still
        INVALID after configure_expressions(), retry PATCH with force_replace.
        """
        logger.info("Validating expression statuses...")
        expressions = expr_manager.load_expressions()
        if not expressions:
            return True  # No expressions to validate

        all_valid = True
        for expr in expressions:
            name = expr["Name"]
            details = expr_manager.get_expression_details(name)
            if not details:
                logger.warning("Expression '%s' not found during validation", name)
                continue

            status = details.get("Status", "INVALID")
            uniq_id = details.get("_uniq_id")

            if status == "VALID":
                logger.info("✅ Expression '%s' Status=VALID", name)
                continue

            # Expression is INVALID — try to fix by (re)setting ExpressionDetails
            logger.warning("⚠ Expression '%s' Status=%s — attempting fix", name, status)
            if uniq_id:
                detail_rows = expr_manager._build_expression_detail_rows(expr)
                if detail_rows:
                    logger.info("🔧 Re-PATCHing %d ExpressionDetails for '%s'", len(detail_rows), name)
                    expr_manager._set_expression_details(
                        uniq_id, name, detail_rows,
                        description=expr.get("Description", name),
                        force_replace=True,
                    )
                    # Re-check status
                    updated = expr_manager.get_expression_details(name)
                    new_status = updated.get("Status", "INVALID") if updated else "UNKNOWN"
                    if new_status == "VALID":
                        logger.info("✅ Expression '%s' fixed → Status=VALID", name)
                    else:
                        logger.warning("⚠ Expression '%s' still %s after fix attempt", name, new_status)
                        all_valid = False
                else:
                    logger.warning("⚠ No detail rows for '%s' — cannot fix", name)
                    all_valid = False
            else:
                logger.warning("⚠ No UniqID for '%s' — cannot fix", name)
                all_valid = False

        return all_valid

    # Execute deployment in dependency order (matches Oracle ICM object hierarchy)
    steps = [
        ("Rate Dimensions", lambda: RateDimensionManager(
            api_client, config_manager, **_mgr_kwargs()
        ).create_rate_dimensions(force=True)),
        ("Rate Tables", lambda: RateTableManager(
            api_client, config_manager, **_mgr_kwargs()
        ).create_rate_tables(force=True)),
        ("Expressions", lambda: expr_manager.configure_expressions(force=True)),
        ("Validate Expressions", _validate_and_fix_expressions),
        ("Performance Measures", lambda: pm_manager.create_performance_measures(
            force=True)),
        ("Plan Components", lambda: PlanComponentManager(
            api_client, config_manager, **_mgr_kwargs(),
            performance_measure_manager=pm_manager,
        ).configure_plan_components(force=True)),
        ("Compensation Plans", lambda: CompensationPlanManager(
            api_client, config_manager, **_mgr_kwargs()
        ).create_compensation_plans_with_components(force=True)),
    ]

    for step_name, step_fn in steps:
        try:
            deploy_log.set_step(step_name)
            logger.info("Deploying: %s", step_name)
            success = step_fn()
            step_entries = [e for e in deploy_log.entries if e["step"] == step_name]
            results["steps"].append({
                "name": step_name,
                "success": bool(success),
                "request_count": len(step_entries),
                "errors": [e for e in step_entries if not e["success"]],
            })
            if not success:
                results["success"] = False
        except Exception as e:
            logger.exception("Deployment step '%s' failed: %s", step_name, e)
            results["steps"].append({
                "name": step_name, "success": False, "error": str(e),
            })
            results["success"] = False

    # Attach detailed log
    results["detailed_log"] = deploy_log.to_dict()

    # Generate post-deployment workbook with Oracle IDs
    try:
        configured_path = generate_configured_workbook(excel_path, deploy_log)
        results["configured_workbook"] = str(configured_path)
        logger.info("Generated configured workbook: %s", configured_path)
    except Exception as e:
        logger.warning("Could not generate configured workbook: %s", e)

    # Clean up normalized workbook
    try:
        normalized_path.unlink(missing_ok=True)
    except Exception:
        pass

    return results
